import os
import time
import shutil
import datetime
import unicodedata
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options

# Prevenir execução múltipla
import sys
if hasattr(sys, '_tangara_running'):
    sys.exit(0)
sys._tangara_running = True

# Configurações de diretórios adaptadas para Windows e Linux
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOAD_DIR = os.path.join(BASE_DIR, 'downloads')
LOG_DIR = os.path.join(BASE_DIR, 'logs')
RELATORIOS_DIR = os.path.join(BASE_DIR, 'relatorios')
ENGENHARIA_DIR = os.path.join(RELATORIOS_DIR, 'engenharia')
SUPRIMENTOS_DIR = os.path.join(RELATORIOS_DIR, 'suprimentos')
SUPRIMENTOS_TANGARA_DIR = os.path.join(SUPRIMENTOS_DIR, 'tangara')
ADMINISTRATIVO_DIR = os.path.join(RELATORIOS_DIR, 'administrativo')

# Credenciais - Usar variáveis de ambiente
EMAIL = os.getenv('TANGARA_EMAIL')
EMAIL_PASSWORD = os.getenv('TANGARA_EMAIL_PASSWORD')

# Criar diretórios se não existirem
for directory in [LOG_DIR, DOWNLOAD_DIR, ENGENHARIA_DIR, SUPRIMENTOS_DIR,
                  SUPRIMENTOS_TANGARA_DIR, ADMINISTRATIVO_DIR]:
    try:
        os.makedirs(directory, exist_ok=True)
    except PermissionError:
        print(f"Aviso: Sem permissão para criar o diretório {directory}")
        # Se falhar, tentamos continuar, talvez o diretório já exista ou será criado manualmente

# Configuração do log
nome_do_arquivo_de_log = f"log_tangara_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.txt"
caminho_do_arquivo_de_log = os.path.join(LOG_DIR, nome_do_arquivo_de_log)

def adicionar_ao_log(mensagem, caminho_log=caminho_do_arquivo_de_log):
    """Adiciona mensagem ao arquivo de log com timestamp"""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    mensagem_formatada = f"{timestamp} - {mensagem}"

    print(mensagem_formatada)  # Sempre imprimir no console primeiro

    try:
        with open(caminho_log, "a", encoding="utf-8") as log_file:
            log_file.write(f"{mensagem_formatada}\n")
    except PermissionError:
        # Apenas ignoramos se não houver permissão de escrita no arquivo de log,
        # pois a mensagem já foi impressa no console (stdout).
        pass

def mostrar_mensagem_conclusao():
    """Mostra mensagem de conclusão"""
    adicionar_ao_log("Programa concluído com sucesso")

def mostrar_mensagem_erro():
    """Mostra mensagem de erro"""
    adicionar_ao_log("Erro na plataforma")

def criar_driver():
    """Cria o driver do Chrome otimizado para Docker"""
    try:
        chrome_options = Options()

        # Opções essenciais para rodar no Docker
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")

        # Opções anti-detecção
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)

        # User agent
        chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

        # Configurações de download
        prefs = {
            "download.default_directory": DOWNLOAD_DIR,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safeBrowse.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)

        driver = webdriver.Chrome(options=chrome_options)

        # Bloqueia os pop-ups do Beamer (novidades/pesquisas) impedindo o
        # carregamento do embed — sem os scripts, o pop-up nunca aparece.
        driver.execute_cdp_cmd('Network.enable', {})
        driver.execute_cdp_cmd('Network.setBlockedURLs', {
            "urls": [
                "*beamer*",
                "*novidades.sienge.com.br*"
            ]
        })

        driver.set_page_load_timeout(60)

        adicionar_ao_log("Driver Chrome criado com sucesso")
        return driver

    except Exception as e:
        adicionar_ao_log(f"Erro ao criar driver: {str(e)}")
        raise

# Aviso de validação inline do Sienge legado (classe + regex). Ex.: 'Informação:
# A unidade construtiva 1 não possui estrutura do planejamento...'. A classe
# spwAlertaAviso restringe ao aviso de verdade (rótulos de form não a usam).
_JS_AVISO_VALIDACAO = r"""
const el = document.querySelector('.spwAlertaAviso, .spwAlertaErro');
if (!el || el.offsetParent === null) return null;
const t = (el.innerText || '').replace(/\s+/g, ' ').trim();
return /\binformação\b|\batenção\b|não possui|só poderá|não há registros|não foi possível/i.test(t) ? t : null;
"""

def texto_aviso_validacao(driver):
    """Texto do aviso de validação visível (spwAlertaAviso/Erro), ou None."""
    try:
        return driver.execute_script(_JS_AVISO_VALIDACAO)
    except Exception:
        return None



class ObraDivergente(Exception):
    """O relatório baixado é de outra obra — arquivo posto em quarentena."""


def _codigo_da_obra_no_relatorio(caminho):
    """Lê o código da obra do cabeçalho de um relatório por obra do Sienge.

    Os relatórios por obra (OrcCom, MedCom, ApIns) trazem nas primeiras linhas
    uma célula '531 - TOM BUENO - IN531 OBRA'. Retorna só o código, ou None se
    não encontrar (relatórios sem obra única não têm esse cabeçalho).
    """
    import re
    from openpyxl import load_workbook
    wb = load_workbook(caminho, read_only=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=20, max_col=8, values_only=True):
            for celula in row:
                if isinstance(celula, str):
                    m = re.match(r"^(\d{2,})\s*-\s*\S", celula.strip())
                    if m:
                        return m.group(1)
        return None
    finally:
        wb.close()


def esperar_download_e_renomear(novo_nome_arquivo, diretorio_destino, wait_time=60, cancelar_se=None, codigo_obra=None):
    """Espera um novo arquivo ser baixado e o renomeia.

    `cancelar_se` é um predicado opcional consultado a cada volta do loop: quando
    fica verdadeiro (ex.: o Sienge exibiu aviso de que o relatório não virá),
    a espera aborta na hora em vez de queimar o timeout inteiro.
    """
    adicionar_ao_log(f"Aguardando download do arquivo '{novo_nome_arquivo}'...")
    inicio = time.time()
    arquivos_antes = set(os.listdir(DOWNLOAD_DIR))

    # O export client-side do SIENGE pode gravar o arquivo no mesmo segundo do
    # clique em 'Exportar' — antes do snapshot acima, que só acontece quando esta
    # função é chamada. O arquivo entrava em 'arquivos_antes' e nunca era visto
    # como novo: na Impulsi, o Painel de Suprimentos esperou os 120s inteiros com
    # o download já concluído no disco. Um arquivo pré-existente com mtime até
    # 30s antes desta chamada é devolvido ao conjunto de candidatos: cada
    # download processado é movido para fora de DOWNLOAD_DIR e os módulos distam
    # minutos, então não há como casar com arquivo de um download anterior.
    recentes_pre_snapshot = set()
    for f in arquivos_antes:
        try:
            if os.path.getmtime(os.path.join(DOWNLOAD_DIR, f)) >= inicio - 30:
                recentes_pre_snapshot.add(f)
        except OSError:
            pass
    arquivos_antes -= recentes_pre_snapshot

    fim_espera = time.time() + wait_time
    arquivo_baixado = None

    while time.time() < fim_espera:
        if cancelar_se is not None and cancelar_se():
            adicionar_ao_log(f"Espera por '{novo_nome_arquivo}' cancelada: a página indicou que o relatório não será gerado.")
            return False

        arquivos_depois = set(os.listdir(DOWNLOAD_DIR))
        novos_arquivos = arquivos_depois - arquivos_antes

        # Filtra temporários: além de .tmp/.crdownload, o Chrome no Linux grava o
        # download num arquivo oculto '.com.google.Chrome.XXXXXX' antes de
        # renomear — ele aparece e some em instantes, e um stat() nele estoura
        # FileNotFoundError (foi o que derrubou o Painel da Inovar).
        arquivos_completos = [f for f in novos_arquivos
                              if not f.endswith(('.tmp', '.crdownload'))
                              and not f.startswith('.')]

        if arquivos_completos:
            try:
                # Pega o arquivo mais recente
                arquivo_baixado = max([os.path.join(DOWNLOAD_DIR, f) for f in arquivos_completos], key=os.path.getctime)
                # Verifica se o arquivo parou de ser modificado
                tamanho_inicial = os.path.getsize(arquivo_baixado)
                time.sleep(2) # Espera 2s para ver se o tamanho muda
                if tamanho_inicial == os.path.getsize(arquivo_baixado):
                    adicionar_ao_log(f"Download concluído: {os.path.basename(arquivo_baixado)}")
                    break # Sai do loop
                else:
                    arquivo_baixado = None # Continua esperando
            except OSError:
                arquivo_baixado = None # arquivo sumiu/renomeou entre o listdir e o stat — tenta de novo

        time.sleep(1) # Pausa antes de verificar novamente

    if arquivo_baixado:
        # Garantia de conteúdo: o mapeamento arquivo->obra é por timing; antes de
        # entregar, confere o código da obra no cabeçalho do xlsx. Divergência vai
        # para quarentena VERIFICAR-<nome> em vez de chegar ao BI com outro nome.
        if codigo_obra and arquivo_baixado.lower().endswith('.xlsx'):
            codigo_no_arquivo = None
            try:
                codigo_no_arquivo = _codigo_da_obra_no_relatorio(arquivo_baixado)
            except Exception as e:
                adicionar_ao_log(f"AVISO: não foi possível conferir a obra do relatório '{novo_nome_arquivo}': {e}")
            if codigo_no_arquivo is None and codigo_obra:
                adicionar_ao_log(f"AVISO: relatório de '{novo_nome_arquivo}' sem cabeçalho de obra — entregue sem conferência.")
            elif codigo_no_arquivo is not None and str(codigo_no_arquivo) != str(codigo_obra):
                quarentena = os.path.join(diretorio_destino, f"VERIFICAR-{novo_nome_arquivo}{os.path.splitext(arquivo_baixado)[1]}")
                if os.path.exists(quarentena):
                    os.remove(quarentena)
                shutil.move(arquivo_baixado, quarentena)
                adicionar_ao_log(
                    f"ERRO: relatório de '{novo_nome_arquivo}' é da obra {codigo_no_arquivo}, "
                    f"esperado {codigo_obra}. Arquivo em quarentena: {quarentena}")
                raise ObraDivergente(f"relatório é da obra {codigo_no_arquivo}, esperado {codigo_obra}")

        extensao = os.path.splitext(arquivo_baixado)[1]
        caminho_destino_final = os.path.join(diretorio_destino, f"{novo_nome_arquivo}{extensao}")

        if os.path.exists(caminho_destino_final):
            try:
                os.remove(caminho_destino_final)
                adicionar_ao_log(f"Arquivo existente removido: {caminho_destino_final}")
            except PermissionError:
                adicionar_ao_log(f"Aviso: Sem permissão para remover arquivo existente: {caminho_destino_final}")

        try:
            shutil.move(arquivo_baixado, caminho_destino_final)
            adicionar_ao_log(f"Arquivo '{os.path.basename(caminho_destino_final)}' salvo em '{diretorio_destino}'")
            return True
        except PermissionError:
            adicionar_ao_log(f"Erro de permissão: Não foi possível mover '{arquivo_baixado}' para '{caminho_destino_final}'")
            return False
    else:
        adicionar_ao_log("Nenhum arquivo novo foi encontrado no tempo esperado.")
        return False


def baixar_relatorio_ou_falhar(driver, novo_nome_arquivo, diretorio_destino, wait_time=120, codigo_obra=None):
    """Espera o download (abortando cedo se o Sienge exibir aviso) e move o
    arquivo. Sem arquivo = exceção: o módulo não pode fechar como OK com o BI
    ficando sem dado novo — antes, o retorno era descartado e um aviso do
    Sienge virava 'sucesso' silencioso com 120s de espera vazia.
    """
    baixou = esperar_download_e_renomear(
        novo_nome_arquivo, diretorio_destino, wait_time=wait_time,
        cancelar_se=lambda: texto_aviso_validacao(driver) is not None,
        codigo_obra=codigo_obra)
    if baixou:
        return
    msg = texto_aviso_validacao(driver)
    if msg:
        adicionar_ao_log(f"AVISO do Sienge: {msg}")
        raise RuntimeError(f"Sienge não gerou '{novo_nome_arquivo}': {msg}")
    raise TimeoutException(f"download de '{novo_nome_arquivo}' não foi encontrado")


def converter_xls_para_xlsx_alternativo(arquivo_entrada):
    """Conversão alternativa de XLS para XLSX usando pandas"""
    try:
        import pandas as pd
        if not os.path.exists(arquivo_entrada):
            raise FileNotFoundError(f"Arquivo não encontrado: {arquivo_entrada}")

        df = pd.read_excel(arquivo_entrada, engine='xlrd')
        arquivo_saida = arquivo_entrada.replace('.xls', '.xlsx')
        df.to_excel(arquivo_saida, index=False, engine='openpyxl')

        adicionar_ao_log(f"Arquivo convertido: {arquivo_saida}")
        if os.path.exists(arquivo_saida):
            try:
                os.remove(arquivo_entrada)
            except PermissionError:
                adicionar_ao_log(f"Aviso: Sem permissão para remover {arquivo_entrada}")

    except Exception as e:
        adicionar_ao_log(f"Aviso: Não foi possível converter XLS para XLSX: {str(e)}")

def fechar_janela(driver, janela_original):
    """Fecha janela popup e retorna à janela original"""
    try:
        WebDriverWait(driver, 5).until(EC.number_of_windows_to_be(2))
        nova_janela = [janela for janela in driver.window_handles if janela != janela_original][0]
        driver.switch_to.window(nova_janela)
        driver.close()
        driver.switch_to.window(janela_original)
        adicionar_ao_log("Janela popup fechada")
    except TimeoutException:
        adicionar_ao_log("Nenhuma janela popup para fechar.")

def marcar_obras(driver, wait, codigo):
    """Marca a obra pelo CÓDIGO (prefixo do texto da linha na consulta).

    O value do radio rowSelect é um índice de linha: muda quando o cadastro de
    obras muda e a ordem DIFERE entre telas (em 14/08/2026 a consulta da 2138
    listava 22001-INCORPORAÇÃO na linha 0 e 33001-OBRA na linha 1, enquanto a
    da 627 só tinha 33001 — o ApIns saía com a INCORPORAÇÃO no lugar da OBRA).
    Falha em achar o código propaga como exceção: seguir sem marcar exportaria
    a obra anteriormente selecionada sob um nome plausível.
    """
    wait.until(EC.element_to_be_clickable((By.XPATH, "//td[img[@title='Abre a consulta']]"))).click()

    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "layerFormConsulta")))

    try:
        elemento = wait.until(EC.presence_of_element_located((By.XPATH,
            f'//input[@type="radio" and @name="rowSelect"]'
            f'[starts-with(normalize-space(ancestor::tr[1]), "{codigo} ")]')))
        elemento.click()
        wait.until(EC.element_to_be_clickable((By.ID, 'pbSelecionar'))).click()
        adicionar_ao_log(f"Obra {codigo} marcada com sucesso")
    finally:
        driver.switch_to.parent_frame()

def configurar_datas_js(driver, id_inicio, id_fim, data_inicio="01/01/2000", data_fim="01/01/2050"):
    """Configura datas usando JavaScript"""
    driver.execute_script(f"""
        document.getElementById('{id_inicio}').value = '{data_inicio}';
        document.getElementById('{id_fim}').value = '{data_fim}';
    """)
    adicionar_ao_log(f"Datas configuradas via JS: {data_inicio} a {data_fim}")

def capturar_screenshot(driver, nome_arquivo=None, pasta_log=None):
    """Salva um screenshot em LOG_DIR (ou em pasta_log, se informada)."""
    if pasta_log is None:
        pasta_log = LOG_DIR

    # Gerar nome do arquivo com timestamp
    if nome_arquivo is None:
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_arquivo = f'screenshot_{timestamp}.png'
    elif not nome_arquivo.endswith('.png'):
        nome_arquivo += '.png'

    # Caminho completo
    caminho_completo = os.path.join(pasta_log, nome_arquivo)

    try:
        # Capturar screenshot
        driver.save_screenshot(caminho_completo)
        print(f"Screenshot salvo em: {caminho_completo}")
        return caminho_completo
    except Exception as e:
        print(f"Erro ao capturar screenshot: {e}")
        return None

def fechar_popups(driver, timeout=3):
    """Fecha banners/avisos comuns (botão de notificação + 'Entendi'). Ignora se não houver.

    Usa espera curta própria em vez do `wait` global de 30s. Com o Beamer
    bloqueado em criar_driver esses banners quase nunca aparecem, e cada espera
    que estourava custava 30s. Quando o banner existe, ele já está no DOM no
    momento em que a página termina de carregar — 3s são suficientes.
    """
    espera = WebDriverWait(driver, timeout)
    for localizador in ((By.XPATH, '/html/body/div[2]/div/div/div[4]/button'),
                        (By.XPATH, "//button[contains(text(), 'Entendi')]")):
        try:
            espera.until(EC.element_to_be_clickable(localizador)).click()
            time.sleep(1)
        except Exception:
            pass

# ------------------------------- Telas novas do Sienge (9.0.4 / MUI) ---------------------------------

def fechar_modais_informativos(driver):
    """Fecha os modais de novidade das telas novas ("Conheça a nova tela..." →
    Fechar; Ajuda Contextual → Entendi). Sem isso o modal intercepta os cliques."""
    for texto in ("Fechar", "FECHAR", "Entendi", "ENTENDI"):
        for botao in driver.find_elements(By.XPATH, f"//button[normalize-space(.)='{texto}']"):
            try:
                if botao.is_displayed():
                    botao.click()
                    adicionar_ao_log(f"Modal informativo fechado ('{texto}').")
                    time.sleep(1)
            except Exception:
                pass

def configurar_datas_mui(driver, wait, name_inicial, name_final,
                         data_inicio="01/01/2000", data_fim="01/01/2050"):
    """Digita o período nos inputs de data das telas novas (MUI).

    O primeiro campo usa espera longa: a tela pode ficar 30s+ em "Carregando..."
    antes de montar o formulário."""
    WebDriverWait(driver, 90).until(EC.element_to_be_clickable(
        (By.XPATH, f"//input[@name='{name_inicial}']")))
    fechar_modais_informativos(driver)
    for name, valor in ((name_inicial, data_inicio), (name_final, data_fim)):
        campo = wait.until(EC.element_to_be_clickable((By.XPATH, f"//input[@name='{name}']")))
        campo.click()
        campo.send_keys(Keys.CONTROL + "a")
        campo.send_keys(valor)
        campo.send_keys(Keys.ENTER)
    adicionar_ao_log(f"Datas configuradas: {data_inicio} a {data_fim}")
    time.sleep(1)

def mostrar_todas_colunas(driver, wait, botao_colunas=True, fechar_painel=False):
    """Abre o menu Colunas (quando existir) e clica em 'Mostrar/Ocultar Todas'.
    Cobre a tela antiga (span exato) e a nova (variações de caixa)."""
    if botao_colunas:
        for rotulo in ("Colunas", "COLUNAS"):
            try:
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, f"//button[normalize-space(.)='{rotulo}']"))).click()
                time.sleep(1)
                break
            except Exception:
                continue
    xpaths = [
        "//span[contains(normalize-space(.),'Mostrar/Ocultar')]",
        "//*[contains(normalize-space(text()),'Mostrar/ocultar todas')]",
        "//input[@name='Mostrar/Ocultar Todas']",
    ]
    for xp in xpaths:
        elementos = [e for e in driver.find_elements(By.XPATH, xp) if e.is_displayed()]
        if elementos:
            driver.execute_script("arguments[0].click();", elementos[0])
            adicionar_ao_log("'Mostrar/Ocultar Todas' aplicado.")
            break
    else:
        adicionar_ao_log("AVISO: 'Mostrar/Ocultar Todas' não localizado.")
    if fechar_painel:
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
        time.sleep(1)

def selecionar_paginacao_todas(driver, wait, rotulo="Todas", rotulos_alternativos=("Todos",), timeout=15):
    """Troca 'Linhas por página' de 25 para 'Todas'/'Todos' (MUI TablePagination).

    O select do MUI abre no 'mousedown' — um click() via JS não o abre (era a
    falha silenciosa da versão anterior); o clique é nativo (ActionChains) com
    fallback na sequência real de eventos de mouse. Só devolve depois de o
    combobox exibir o novo rótulo; senão levanta TimeoutException para o
    chamador tentar de novo. O rótulo varia por tela: 'Todas' no Painel de
    Suprimentos, 'Todos' no Cadastro de Contratos.
    """
    rotulos = (rotulo, *rotulos_alternativos)
    espera = WebDriverWait(driver, timeout)

    # Overlays que interceptam o clique no rodapé (popover 'Entendi', 'Fechar')
    for xp in (
        "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'entendi')]",
        "//button[@aria-label='Close' or @aria-label='Fechar']",
    ):
        for botao in driver.find_elements(By.XPATH, xp):
            try:
                if botao.is_displayed():
                    botao.click()
                    time.sleep(0.5)
            except Exception:
                pass

    # É um div[role=combobox], NÃO um <select>
    combo_xpath = "//div[@role='combobox' and contains(@class,'MuiTablePagination-select')]"
    dropdown = espera.until(EC.presence_of_element_located((By.XPATH, combo_xpath)))
    atual = dropdown.text.strip()
    if atual in rotulos:
        adicionar_ao_log(f"Paginação já está em '{atual}'.")
        return atual
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", dropdown)
    time.sleep(0.5)

    menu_xpath = "//ul[@role='listbox']"
    try:
        ActionChains(driver).move_to_element(dropdown).click().perform()
    except Exception:
        pass
    if not driver.find_elements(By.XPATH, menu_xpath):
        driver.execute_script("""
            const el = arguments[0];
            const o = {bubbles:true, cancelable:true, view:window};
            ['pointerdown','mousedown','mouseup','click'].forEach(t =>
                el.dispatchEvent(new MouseEvent(t, o)));
        """, dropdown)
    espera.until(EC.presence_of_element_located((By.XPATH, menu_xpath)))
    time.sleep(0.5)

    # O data-value é dinâmico (= total de linhas), NÃO '-1'.
    condicao = " or ".join(f"normalize-space(.)='{r}'" for r in rotulos)
    try:
        opcao = espera.until(EC.element_to_be_clickable((By.XPATH, f"{menu_xpath}//li[{condicao}]")))
    except TimeoutException:
        # 'Todas' fica por último no menu
        opcao = driver.find_elements(By.XPATH, f"{menu_xpath}//li[@role='option']")[-1]
    escolhido = opcao.text.strip()
    ActionChains(driver).move_to_element(opcao).click().perform()

    espera.until(lambda d: d.find_element(By.XPATH, combo_xpath).text.strip() in rotulos)
    adicionar_ao_log(f"Paginação alterada para '{escolhido}'.")
    return escolhido


# Status do MuiDataGrid (telas novas do Sienge que exportam via 'Gerar Relatório'):
#  - carregadas: linhas já materializadas no modelo (altura do virtualScroller /
#    altura de uma linha). Com 'Todas' o grid faz UM fetch longo; até a resposta
#    chegar a altura fica ~= 1 página;
#  - total: total exibido no rodapé de paginação ("1–127 de 127");
#  - carregando: spinner/barra de loading do grid visível.
_JS_STATUS_DATAGRID = r"""
const vs = document.querySelector('.MuiDataGrid-virtualScroller');
const row = document.querySelector('.MuiDataGrid-row');
const foot = document.querySelector('.MuiTablePagination-displayedRows');
const rh = (row && row.offsetHeight) ? row.offsetHeight : 36;
const carregadas = vs ? Math.round((vs.scrollHeight || 0) / rh) : 0;
let total = null;
if (foot) {
    const m = foot.textContent.replace(/[.\s ]/g, '').match(/de([0-9]+)/i);
    if (m) total = parseInt(m[1], 10);
}
const spinner = document.querySelector(
    '.MuiDataGrid-root [role="progressbar"], .MuiDataGrid-loadingOverlay, '
    + '.MuiDataGrid-overlay .MuiCircularProgress-root'
);
const carregando = !!(spinner && spinner.offsetParent !== null);
return {carregadas: carregadas, total: total, carregando: carregando};
"""


def esperar_datagrid_carregar_todas(driver, timeout=300, estagnado=20):
    """Após a paginação 'Todas', espera o grid materializar TODAS as linhas.

    Pronto quando o spinner some E as linhas materializadas alcançam o total do
    rodapé. Se o grid fica parado (sem spinner, contagem inalterada) por
    `estagnado` segundos abaixo do total, a paginação não pegou: devolve cedo
    para o chamador tentar de novo em vez de queimar o timeout inteiro (eram
    300s mortos em 26/2305). Retorna (carregadas, total); o chamador decide.
    """
    deadline = time.time() + timeout
    carregadas = total = None
    proximo_log = 0.0
    ultimo_valor, parado_desde = None, time.time()
    while time.time() < deadline:
        try:
            status = driver.execute_script(_JS_STATUS_DATAGRID)
            carregadas = status.get("carregadas")
            total = status.get("total")
            carregando = status.get("carregando")
        except Exception:
            carregadas = total = None
            carregando = True
        if total and carregadas and carregadas >= total * 0.99 and not carregando:
            adicionar_ao_log(f"DataGrid carregou {carregadas}/{total} linhas.")
            return carregadas, total
        if carregando or carregadas != ultimo_valor:
            ultimo_valor, parado_desde = carregadas, time.time()
        elif time.time() - parado_desde >= estagnado:
            adicionar_ao_log(f"DataGrid parado em {carregadas}/{total} linhas há {estagnado}s sem carregar.")
            return carregadas, total
        if time.time() >= proximo_log:
            adicionar_ao_log(f"Aguardando DataGrid carregar linhas... {carregadas}/{total} "
                             f"(carregando={carregando})")
            proximo_log = time.time() + 15
        time.sleep(1)
    adicionar_ao_log(f"AVISO: timeout ({timeout}s) esperando o DataGrid "
                     f"(materializadas={carregadas}, total={total}).")
    return carregadas, total


class GridIncompleto(RuntimeError):
    """O grid não materializou todas as linhas — o export sairia truncado."""


def garantir_grid_completo(driver, wait, tentativas=2, timeout=300):
    """Paginação 25 → 'Todas' e espera real das linhas, com nova tentativa.

    O export do Sienge 9.0.4 é client-side: 'Gerar Relatório' entrega só o que
    está materializado no grid. Sem esta etapa o arquivo sai com a 1ª página
    (25 linhas + cabeçalho). Se depois de `tentativas` o grid seguir incompleto,
    levanta GridIncompleto: o módulo falha explicitamente em vez de o BI receber
    um relatório truncado com cara de sucesso.
    """
    carregadas = total = None
    for tentativa in range(1, tentativas + 1):
        try:
            selecionar_paginacao_todas(driver, wait)
        except Exception as e:
            adicionar_ao_log(f"AVISO: falha ao trocar paginação para 'Todas' "
                             f"(tentativa {tentativa}/{tentativas}): {e}")
        carregadas, total = esperar_datagrid_carregar_todas(driver, timeout=timeout)
        if total and carregadas and carregadas >= total * 0.99:
            return carregadas, total
        adicionar_ao_log(f"AVISO: grid incompleto ({carregadas}/{total} linhas) "
                         f"na tentativa {tentativa}/{tentativas}.")
    raise GridIncompleto(f"grid materializou {carregadas} de {total} linhas — o export sairia truncado")

def esperar_datagrid_pronto(driver, timeout=60):
    """Espera a primeira carga do MuiDataGrid (spinner/overlay ausente).

    Diferente de esperar_datagrid_carregar_todas: serve para o pós-Consultar,
    quando a paginação ainda está em 25 e 'todas as linhas' nunca materializam.
    """
    def _carregando():
        try:
            return driver.execute_script(
                "const g=document.querySelector('.MuiDataGrid-root');"
                "if(!g) return true;"
                "return !!g.querySelector('.MuiLinearProgress-root')"
                " || g.getAttribute('aria-busy')==='true'"
                " || !!g.querySelector('.MuiDataGrid-loadingOverlay');")
        except Exception:
            return True
    fim = time.time() + timeout
    while time.time() < fim:
        if not _carregando():
            time.sleep(1)
            return True
        time.sleep(0.5)
    adicionar_ao_log(f"AVISO: DataGrid não ficou pronto em {timeout}s após a consulta.")
    return False


def exportar_excel_mui(driver, wait):
    """Gerar Relatório → formato excel → Exportar (dispara o download).

    O combobox e o botão são procurados DENTRO do diálogo: '//div[@role=...
    presentation]//div[@role=combobox]' também casa com o select de paginação
    do grid, que fica atrás do modal (clique interceptado)."""
    wait.until(EC.element_to_be_clickable((By.XPATH,
        "//button[normalize-space(.)='Gerar Relatório' or normalize-space(.)='GERAR RELATÓRIO']"))).click()
    adicionar_ao_log("Botão 'Gerar Relatório' clicado.")
    time.sleep(1)

    # Varre apenas dialogs VISÍVEIS: pode haver .MuiDialog-container ocultos no
    # DOM e um XPath first-match trava neles. O click no combobox é nativo — o
    # Select do MUI abre no mousedown, que um click() via JS não dispara.
    def _no_dialogo_visivel(seletor):
        return driver.execute_script("""
            const seletor = arguments[0];
            for (const d of document.querySelectorAll('.MuiDialog-container, [role="dialog"]')) {
                if (d.offsetParent === null) continue;
                for (const el of d.querySelectorAll(seletor)) {
                    if (el.offsetParent !== null) return el;
                }
            }
            return null;
        """, seletor)

    fim = time.time() + 20
    combobox = None
    while time.time() < fim and combobox is None:
        combobox = _no_dialogo_visivel("[role='combobox'], .MuiSelect-select")
        if combobox is None:
            time.sleep(0.5)
    if combobox is None:
        raise TimeoutException("combobox de formato do relatório não localizado")
    try:
        combobox.click()
    except Exception:
        driver.execute_script(
            "arguments[0].dispatchEvent(new MouseEvent('mousedown', {bubbles: true}));", combobox)

    opcao_excel = wait.until(EC.presence_of_element_located((By.XPATH, "//li[@data-value='excel']")))
    driver.execute_script("arguments[0].click();", opcao_excel)
    time.sleep(1)

    fim = time.time() + 20
    exportar = None
    while time.time() < fim and exportar is None:
        exportar = driver.execute_script("""
            for (const d of document.querySelectorAll('.MuiDialog-container, [role="dialog"]')) {
                if (d.offsetParent === null) continue;
                for (const b of d.querySelectorAll('button')) {
                    if (/^exportar$/i.test((b.textContent || '').trim())) return b;
                }
            }
            return null;
        """)
        if exportar is None:
            time.sleep(0.5)
    if exportar is None:
        raise TimeoutException("botão 'Exportar' não localizado no modal")
    driver.execute_script("arguments[0].click();", exportar)
    adicionar_ao_log("Botão 'Exportar' clicado.")


# -----------------------------------------------------------------------------------------------------------------------------------
# ------------------------------------------------- RESILIÊNCIA E ORQUESTRAÇÃO -----------------------------------------------------
# -----------------------------------------------------------------------------------------------------------------------------------


class ModuloSemAcesso(Exception):
    """O usuário logado não tem permissão para o módulo — não é falha do robô."""


# Frases (minúsculas) que o SIENGE exibe quando o usuário não tem permissão.
# Mantidas restritas de propósito: um falso positivo aqui pularia um módulo bom.
FRASES_SEM_ACESSO = (
    "não tem permissão",
    "nao tem permissao",
    "não possui permissão",
    "nao possui permissao",
    "sem permissão de acesso",
    "sem permissao de acesso",
    "permissão insuficiente",
    "permissao insuficiente",
    "acesso negado",
    "acesso não autorizado",
    "acesso nao autorizado",
    "usuário não autorizado",
    "usuario nao autorizado",
    "não tem autorização",
    "nao tem autorizacao",
)

STATUS_OK = "OK"
STATUS_FALHOU = "FALHOU"
STATUS_SEM_ACESSO = "SEM ACESSO"


def _texto_da_pagina(driver):
    """Texto visível da página e dos iframes de mesma origem, em minúsculas."""
    return driver.execute_script("""
        let t = document.body ? document.body.innerText : '';
        for (const f of document.querySelectorAll('iframe')) {
            try { t += '\\n' + f.contentDocument.body.innerText; } catch (e) {}
        }
        return t.toLowerCase();
    """) or ""


def _frase_de_bloqueio(texto):
    for frase in FRASES_SEM_ACESSO:
        if frase in texto:
            return frase
    return None


def verificar_acesso(driver, nome_modulo, espera=3):
    """Detecta cedo a página de 'sem permissão' do SIENGE.

    Sem esta checagem, um módulo sem acesso só falha depois de vários
    WebDriverWait estourando (30s cada). Levanta ModuloSemAcesso para que
    executar_modulo registre o motivo real e siga para o próximo módulo.

    O SIENGE é uma SPA com rotas em hash: um driver.get() nem sempre recarrega a
    página, então o texto logo após a navegação pode ainda ser o do módulo
    anterior. Por isso a frase de bloqueio precisa aparecer em duas leituras
    separadas — evita pular um módulo bom por causa de conteúdo velho na tela.
    """
    if espera:
        time.sleep(espera)

    try:
        frase = _frase_de_bloqueio(_texto_da_pagina(driver))
        if frase is None:
            return

        time.sleep(2)
        frase = _frase_de_bloqueio(_texto_da_pagina(driver))
        if frase is None:
            return
    except Exception as e:
        adicionar_ao_log(f"AVISO: não foi possível ler a página para checar acesso: {e}")
        return

    raise ModuloSemAcesso(f"{nome_modulo}: página retornou \"{frase}\"")


def slug_ascii(texto):
    """Converte um nome de módulo em nome de arquivo seguro.

    Os nomes de módulo têm acentos e parênteses ('Apropriações de Insumos
    (Engenharia)'). Os screenshots vão para uma pasta que é lida por fora,
    então o nome é reduzido a ASCII: acento vira letra simples, o resto vira '_'.
    """
    sem_acento = (
        unicodedata.normalize("NFKD", texto)
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    slug = "".join(c if c.isalnum() else "_" for c in sem_acento.lower())
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug.strip("_")


def sanear_estado(driver, janela_original=None):
    """Devolve o driver a um estado neutro entre módulos.

    Um módulo que quebra no meio costuma deixar o driver dentro de um iframe,
    com uma aba de relatório aberta ou com um modal na frente. Sem esta limpeza
    o módulo seguinte falharia por herança e o isolamento não teria efeito.
    """
    try:
        if janela_original and janela_original in driver.window_handles:
            for handle in list(driver.window_handles):
                if handle != janela_original:
                    driver.switch_to.window(handle)
                    driver.close()
            driver.switch_to.window(janela_original)
    except Exception as e:
        adicionar_ao_log(f"AVISO: falha ao fechar janelas extras: {e}")

    try:
        driver.switch_to.default_content()
    except Exception:
        pass

    try:
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
    except Exception:
        pass


def executar_modulo(driver, funcao_modulo, nome_modulo, wait_local, janela_original=None):
    """Executa um módulo isolando suas falhas do resto da automação.

    Retorna STATUS_OK, STATUS_SEM_ACESSO ou STATUS_FALHOU — nunca propaga a
    exceção, para que os demais módulos continuem rodando.
    """
    adicionar_ao_log(f"\n--- INICIANDO MÓDULO: {nome_modulo} ---")
    sanear_estado(driver, janela_original)

    try:
        funcao_modulo(driver, wait_local)
        adicionar_ao_log(f"--- MÓDULO '{nome_modulo}' FINALIZADO COM SUCESSO ---")
        return STATUS_OK

    except ModuloSemAcesso as e:
        adicionar_ao_log(f"### MÓDULO IGNORADO — SEM ACESSO: {e}")
        adicionar_ao_log("### A execução continuará com o próximo módulo.")
        return STATUS_SEM_ACESSO

    except Exception as e:
        nome_screenshot = f"erro_{slug_ascii(nome_modulo)}"
        adicionar_ao_log("!")
        adicionar_ao_log(f"!!! ERRO AO EXECUTAR O MÓDULO: {nome_modulo} !!!")
        adicionar_ao_log(f"!!! ERRO: {e}")
        adicionar_ao_log(
            f"!!! A execução continuará com o próximo módulo. Screenshot: {nome_screenshot}.png"
        )
        adicionar_ao_log("!")
        capturar_screenshot(driver, nome_screenshot)
        return STATUS_FALHOU

    finally:
        sanear_estado(driver, janela_original)


def resumo_execucao(resultados):
    """Imprime no log o placar final por módulo."""
    if not resultados:
        return

    adicionar_ao_log("\n========== RESUMO DA EXECUÇÃO ==========")
    for nome, status in resultados:
        adicionar_ao_log(f"[{status}] {nome}")

    ok = sum(1 for _, status in resultados if status == STATUS_OK)
    sem_acesso = sum(1 for _, status in resultados if status == STATUS_SEM_ACESSO)
    falhou = sum(1 for _, status in resultados if status == STATUS_FALHOU)
    adicionar_ao_log(
        f"{ok}/{len(resultados)} módulos com sucesso "
        f"({sem_acesso} sem acesso, {falhou} com erro)."
    )
    adicionar_ao_log("========================================")


# -----------------------------------------------------------------------------------------------------------------------------------
# ------------------------------------------------- MÓDULOS ------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------------------------------------------

def executar_login(driver, wait):
    """Login no SIENGE. Etapa crítica: se falhar, nenhum módulo pode rodar.

    Retorna o handle da janela original, usado por sanear_estado() para
    fechar abas de relatório deixadas para trás por módulos que quebraram.
    """
    adicionar_ao_log("Acessando página do SIENGE TANGARA...")
    driver.get("https://guzattizompero.sienge.com.br/sienge/")

    wait.until(EC.element_to_be_clickable((By.ID, "btnEntrarComSiengeID"))).click()
    adicionar_ao_log("Botão de login clicado")

    adicionar_ao_log("Verificando tela de login adicional...")
    email_input_ms = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//input[@name='email']"))
    )
    email_input_ms.send_keys(EMAIL)
    email_input_ms.send_keys(Keys.ENTER)
    adicionar_ao_log("E-mail inserido.")

    password_input_ms = wait.until(
        EC.visibility_of_element_located((By.XPATH, "//input[@type='password']"))
    )
    password_input_ms.send_keys(EMAIL_PASSWORD)
    password_input_ms.send_keys(Keys.ENTER)
    adicionar_ao_log("Senha inserida na tela.")

    try:
        # Aguarda até 5 segundos para o alerta aparecer na tela (caso o usuário já esteja logado)
        WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.XPATH, "//div[contains(@class,'spwAlertaAviso')]"))
        )
        driver.find_element(By.CLASS_NAME, "Button-prim").click()
        adicionar_ao_log("Alerta de aviso fechado.")
    except Exception:
        adicionar_ao_log("Nenhum alerta de aviso encontrado.")

    # Espera o carregamento da página principal pós-login
    wait.until(EC.title_contains("Sienge"))
    adicionar_ao_log("Login realizado com sucesso, página principal carregada.")

    return driver.current_window_handle


def modulo_cadastro_contratos(driver, wait):
    """Cadastro de Contratos (Suprimentos).

    O Sienge trocou esta tela em 2026 (9.0.4): modais de novidade na entrada,
    datas por input[name], painel de colunas novo e export client-side (só as
    linhas materializadas no grid). Fluxo: fechar modais → datas 2000→2050 →
    mostrar todas as colunas → Consultar → paginação 'Todos' → esperar a grid
    materializar tudo (ou falhar) → Gerar Relatório → excel.
    """
    adicionar_ao_log("Iniciando extração de Cadastro de Contratos...")
    driver.get("https://guzattizompero.sienge.com.br/sienge/8/index.html#/suprimentos/contratos-e-medicoes/contratos/cadastros")
    time.sleep(3)
    verificar_acesso(driver, "Cadastro de Contratos", espera=0)

    # Use ActionChains to send ESCAPE key to the active element
    ActionChains(driver).send_keys(Keys.ESCAPE).perform()
    fechar_popups(driver)
    fechar_modais_informativos(driver)

    # Configurar relatório
    configurar_datas_mui(driver, wait, "dtContratoInicial", "dtContratoFinal")
    mostrar_todas_colunas(driver, wait, botao_colunas=True, fechar_painel=True)

    btConsultar = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space(.)='Consultar' or normalize-space(.)='CONSULTAR']")))
    driver.execute_script("arguments[0].click();", btConsultar)
    esperar_datagrid_pronto(driver)

    # O export é client-side (só as linhas materializadas): paginação 'Todos' +
    # espera real, com nova tentativa. Sem completar, o módulo FALHA — antes o
    # erro virava AVISO e o BI recebia um arquivo com uma página só (26 linhas).
    garantir_grid_completo(driver, wait)
    time.sleep(1)

    capturar_screenshot(driver, "cadastro_contratos_grid")

    # Exportar para Excel
    exportar_excel_mui(driver, wait)

    baixar_relatorio_ou_falhar(driver, "cadastro de contratos", ADMINISTRATIVO_DIR, wait_time=120)


def modulo_analitico_apropriacoes(driver, wait):
    """Analítico de Apropriações por Obra — Emissão e Vencimento (Engenharia)."""
    adicionar_ao_log("Acessando Analítico de Apropriações por Obra...")
    driver.get("https://guzattizompero.sienge.com.br/sienge/8/index.html#/common/page/588")
    verificar_acesso(driver, "Analítico de Apropriações por Obra")
    fechar_popups(driver)

    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'iFramePage')))

    Select(wait.until(EC.visibility_of_element_located((By.NAME, 'analise.selecao')))).select_by_value("emissao")

    wait.until(EC.element_to_be_clickable((By.XPATH, "//td[img[@title='Abre a consulta']]"))).click()
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "layerFormConsulta")))
    wait.until(EC.element_to_be_clickable((By.ID, 'pbMarcarTodos'))).click()
    wait.until(EC.element_to_be_clickable((By.ID, 'pbSelecionar'))).click()
    driver.switch_to.parent_frame()

    # IMPORTANTE: A marcação da Obra realiza um AJAX Refresh no formulário do Sienge, restabelecendo as datas para os padrões do projeto (ex: 2025)!
    # Sendo assim, é OBRIGATÓRIO que o preenchimento da data aconteça DEPOIS da seleção da Obra para que o valor não seja perdido.
    time.sleep(2) # Pausa para aguardar o AJAX repopular os campos do Sienge pós formConsulta
    configurar_datas_js(driver, "analise.periodoInicio", "analise.periodoFim")

    wait.until(EC.element_to_be_clickable((By.ID, 'analise.imprimirObservacaoTitulo'))).click()
    wait.until(EC.element_to_be_clickable((By.ID, 'analise.imprimirDadosEmColunasNaoMescladas'))).click()
    capturar_screenshot(driver, "analitico_de_apropriacoes.png", LOG_DIR)

    wait.until(EC.element_to_be_clickable((By.ID, 'visualizarButton'))).click()
    baixar_relatorio_ou_falhar(driver, "Analítico de Apropriações por Obra EMISSAO - HERANZA - TANGARA", ENGENHARIA_DIR, wait_time=120)

    # Gerar relatório VENCIMENTO
    Select(driver.find_element(By.NAME, 'analise.selecao')).select_by_value("vencimento")
    capturar_screenshot(driver, "analitico_de_apropriacoes_vencimento.png", LOG_DIR)
    wait.until(EC.element_to_be_clickable((By.ID, 'visualizarButton'))).click()
    baixar_relatorio_ou_falhar(driver, "Analítico de Apropriações por Obra VENCIMENTO - HERANZA - TANGARA", ENGENHARIA_DIR, wait_time=120)

    driver.switch_to.default_content()


def modulo_orcado_comprometido(driver, wait):
    """Comparativo Orçado x Comprometido (Engenharia)."""
    adicionar_ao_log("Acessando Comparativo Orçado x Comprometido...")
    driver.get("https://guzattizompero.sienge.com.br/sienge/8/index.html#/common/page/627")
    verificar_acesso(driver, "Orçado x Comprometido")
    fechar_popups(driver)

    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'iFramePage')))

    marcar_obras(driver, wait, "33001")

    configurar_datas_js(driver, "analise.periodoInicio", "analise.periodoFim")

    Select(driver.find_element(By.NAME, 'analise.selecao')).select_by_value("emissao")
    Select(driver.find_element(By.ID, "analise.nivelDetalhamento")).select_by_value("4")
    Select(driver.find_element(By.ID, "analise.bdi")).select_by_value("N")
    Select(driver.find_element(By.ID, "analise.leiSocial")).select_by_value("N")

    for checkbox_id in ['analise.consDocPrev',
                        'analise.impPercRealiItensOrc',
                        'analise.impVlEstoqAtualObra',
                        'analise.impVlEstoqServico',
                        'analise.apreDifCompOrcEmVl',
                        'analise.ocultarRegistroSemMovimentacao'
                        ]:
        wait.until(EC.element_to_be_clickable((By.ID, checkbox_id))).click()

    wait.until(EC.element_to_be_clickable((By.ID, 'btOpcoesRelatorio'))).click()
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "layerFormConsulta")))
    Select(driver.find_element(By.ID, 'formatoSaidaDocumento')).select_by_value("XLSX")
    wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/table/tbody/tr[3]/td/table/tbody/tr/td[1]/span[1]/span/input'))).click()
    driver.switch_to.parent_frame()

    capturar_screenshot(driver, "comparativo_orcado_x_comprometido.png", LOG_DIR)
    wait.until(EC.element_to_be_clickable((By.ID, 'visualizarButton'))).click()
    baixar_relatorio_ou_falhar(driver, "OrcCom-HERANZA - TANGARA", ENGENHARIA_DIR, wait_time=120, codigo_obra="33001")
    driver.switch_to.default_content()


def modulo_medido_comprometido(driver, wait):
    """Comparativo Medido x Comprometido (Engenharia)."""
    adicionar_ao_log("Acessando Comparativo Medido x Comprometido...")
    driver.get("https://guzattizompero.sienge.com.br/sienge/8/index.html#/common/page/623")
    verificar_acesso(driver, "Medido x Comprometido")
    fechar_popups(driver)

    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'iFramePage')))

    marcar_obras(driver, wait, "33001")

    configurar_datas_js(driver, "analise.periodoInicio", "analise.periodoFim")
    Select(driver.find_element(By.NAME, 'analise.selecao')).select_by_value("emissao")
    Select(driver.find_element(By.ID, "analise.nivelDetalhamento")).select_by_value("0")
    Select(driver.find_element(By.ID, "analise.bdi")).select_by_value("N")
    Select(driver.find_element(By.ID, "analise.leiSocial")).select_by_value("N")

    for checkbox_id in ['analise.consDocPrev',
                        'analise.impPercRealiItensOrc',
                        'analise.impVlEstoqAtualObra',
                        'analise.impVlEstoqTarefa',
                        'analise.impCodOrc'
                        ]:
        wait.until(EC.element_to_be_clickable((By.ID, checkbox_id))).click()

    capturar_screenshot(driver, "comparativo_medido_x_comprometido.png", LOG_DIR)

    wait.until(EC.element_to_be_clickable((By.ID, 'visualizarButton'))).click()
    baixar_relatorio_ou_falhar(driver, "MedCom-HERANZA - TANGARA", ENGENHARIA_DIR, wait_time=120, codigo_obra="33001")
    driver.switch_to.default_content()


def modulo_apropriacoes_insumos(driver, wait):
    """Apropriações de Insumos (Engenharia)."""
    adicionar_ao_log("Acessando Apropriações de Insumos...")
    driver.get("https://guzattizompero.sienge.com.br/sienge/8/index.html#/common/page/2138")
    verificar_acesso(driver, "Apropriações de Insumos")
    fechar_popups(driver)

    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'iFramePage')))
    marcar_obras(driver, wait, "33001")
    configurar_datas_js(driver, "filter.dataInicialPeriodo", "filter.dataFinalPeriodo")
    Select(driver.find_element(By.ID, 'tpBDI')).select_by_value("N")
    Select(driver.find_element(By.ID, 'tpEncargosSociais')).select_by_value("N")

    wait.until(EC.element_to_be_clickable((By.ID, "filter.imprimirSemQuantidades"))).click()
    wait.until(EC.element_to_be_clickable((By.ID, "imprimirPedidosPendentes"))).click()
    wait.until(EC.element_to_be_clickable((By.ID, "imprimirContratosPendentes"))).click()

    wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='submit' and @name='btFiltrar']"))).click()
    baixar_relatorio_ou_falhar(driver, "ApIns-HERANZA - TANGARA", ENGENHARIA_DIR, wait_time=120, codigo_obra="33001")
    driver.switch_to.default_content()


def modulo_painel_suprimentos(driver, wait):
    """Painel de Suprimentos (Suprimentos)."""
    adicionar_ao_log("Acessando Painel de Suprimentos...")
    driver.get("https://guzattizompero.sienge.com.br/sienge/8/index.html#/suprimentos/compras/painel-de-compras")
    time.sleep(2)
    verificar_acesso(driver, "Painel de Suprimentos", espera=0)
    fechar_popups(driver)
    fechar_modais_informativos(driver)

    driver.switch_to.default_content()

    # Configurar datas
    data_inicial = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@name='dataInicial']")))
    data_inicial.click()
    data_inicial.send_keys(Keys.CONTROL + "a")
    data_inicial.send_keys("01/01/2000")
    data_inicial.send_keys(Keys.ENTER)

    data_final = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@name='dataFinal']")))
    data_final.click()
    data_final.send_keys(Keys.CONTROL + "a")
    data_final.send_keys("01/01/2050")
    data_final.send_keys(Keys.ENTER)
    time.sleep(1)

    # Dispara consulta para listar todos os registros no período
    try:
        consultar_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space(.)='Consultar' or normalize-space(.)='CONSULTAR']")))
        driver.execute_script("arguments[0].click();", consultar_btn)
        adicionar_ao_log("Botão 'Consultar' clicado.")
    except Exception as e:
        adicionar_ao_log(f"Aviso: botão 'Consultar' não encontrado ou já acionado: {e}")

    # Aguarda a primeira carga da consulta. NÃO usar esperar_datagrid_carregar_todas
    # aqui: antes da paginação 'Todas' o grid materializa só as 25 linhas da página
    # e a condição 'carregadas >= total' nunca fecha — eram 300s de espera morta
    # em toda run. A espera de todas as linhas acontece após a troca de paginação.
    esperar_datagrid_pronto(driver)
    capturar_screenshot(driver, "painel_suprimentos_consulta")

    # Mostrar/Ocultar Todas — Cód. Obra/Insumo/Grupo vêm DESMARCADOS por padrão.
    # O controle fica dentro do popover "Exibir seletor de colunas" (botão de ícone,
    # sem texto), por isso os antigos gatilhos 'COLUNAS'/'FILTROS' nunca o encontravam.
    try:
        def _clicar_robusto(el):
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            try:
                ActionChains(driver).move_to_element(el).click().perform()
            except Exception:
                driver.execute_script("arguments[0].click();", el)

        # Abre o seletor de colunas (botão de ícone, sem texto)
        col_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[@aria-label='Exibir seletor de colunas']")))
        _clicar_robusto(col_btn)
        time.sleep(1)

        # Checkbox-mestre 'Mostrar/Ocultar Todas' (label MuiFormControlLabel)
        master = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//label[.//span[normalize-space(.)='Mostrar/Ocultar Todas']]")))
        master_input = master.find_element(By.XPATH, ".//input[@type='checkbox']")

        # Clica até ficar 'checked' e não-indeterminado (= todas as colunas visíveis).
        # Estado inicial costuma ser indeterminado → 1 clique marca todas.
        for _ in range(3):
            checked = bool(master_input.get_property('checked'))
            indeterminate = bool(driver.execute_script("return arguments[0].indeterminate;", master_input))
            if checked and not indeterminate:
                break
            _clicar_robusto(master)
            time.sleep(0.7)
        adicionar_ao_log("'Mostrar/Ocultar Todas' marcado (todas as colunas visíveis).")

        # Fecha o popover de colunas
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        time.sleep(1)
    except Exception as e:
        adicionar_ao_log(f"AVISO: erro em 'Mostrar/Ocultar Todas': {e}")
    time.sleep(2)

    # Troca paginação 25 → Todas + espera real das linhas materializarem.
    # CRÍTICO: trocar para 'Todas' dispara uma busca no servidor de TODAS as linhas.
    # O export do SIENGE é client-side ("exporta o que está visível na tabela"),
    # então exportar antes do fim da busca traz só as 25 linhas da 1ª página.
    try:
        selecionar_paginacao_todas(driver, wait, rotulo="Todas", rotulos_alternativos=("Todos",))
        esperar_datagrid_carregar_todas(driver)
        time.sleep(1)  # margem extra para o modelo de dados assentar
        capturar_screenshot(driver, "painel_suprimentos_todas")
    except Exception as e:
        adicionar_ao_log(f"AVISO: falha ao trocar paginação 25 → Todas: {e}")

    # Gera relatório em Excel (ancorado no diálogo — ver exportar_excel_mui)
    exportar_excel_mui(driver, wait)
    capturar_screenshot(driver, "painel_suprimentos_modal_excel")

    # Sem arquivo = módulo falhou; não pode fechar como OK com o BI sem dado novo.
    baixar_relatorio_ou_falhar(driver, "PAINEL DE SUPRIMENTOS - TANGARA", SUPRIMENTOS_TANGARA_DIR, wait_time=120)


# -----------------------------------------------------------------------------------------------------------------------------------
# ------------------------------------------------- MAIN ----------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------------------------------------------


# Módulos disponíveis: (id, função, nome no log). O id é o que a variável
# MODULOS e a Central de Automações usam para escolher o que rodar.
MODULOS = [
    ("cadastro_contratos", modulo_cadastro_contratos, "Cadastro de Contratos (Suprimentos)"),
    ("analitico_apropriacoes", modulo_analitico_apropriacoes, "Analítico de Apropriações por Obra (Engenharia)"),
    ("orcado_comprometido", modulo_orcado_comprometido, "Orçado x Comprometido (Engenharia)"),
    ("medido_comprometido", modulo_medido_comprometido, "Medido x Comprometido (Engenharia)"),
    ("apropriacoes_insumos", modulo_apropriacoes_insumos, "Apropriações de Insumos (Engenharia)"),
    ("painel_suprimentos", modulo_painel_suprimentos, "Painel de Suprimentos (Suprimentos)"),
]


def selecionar_modulos(modulos, selecao):
    """Filtra `modulos` [(id, funcao, nome)] pela variável de ambiente MODULOS ("a,b").

    Vazia ou ausente = todos. Id desconhecido aborta antes de abrir o navegador,
    listando os válidos. A ordem é sempre a da lista, não a da seleção.
    """
    ids = [s.strip() for s in (selecao or "").split(",") if s.strip()]
    if not ids:
        return list(modulos)
    validos = [m[0] for m in modulos]
    desconhecidos = [i for i in ids if i not in validos]
    if desconhecidos:
        raise ValueError(
            f"MODULOS desconhecidos: {', '.join(desconhecidos)}. Válidos: {', '.join(validos)}")
    return [m for m in modulos if m[0] in ids]


def main():
    """Roda o login (crítico) e depois cada módulo de forma isolada.

    Falha de um módulo — inclusive falta de permissão do usuário — é registrada
    e não interrompe os demais. Só erros de setup/login abortam a automação.
    """
    driver = None
    resultados = []

    try:
        adicionar_ao_log("========================================")
        adicionar_ao_log("===== INICIANDO AUTOMAÇÃO TANGARA ======")
        adicionar_ao_log("========================================")

        selecionados = selecionar_modulos(MODULOS, os.environ.get("MODULOS"))
        if len(selecionados) != len(MODULOS):
            adicionar_ao_log(f"MODULOS={os.environ.get('MODULOS')}: executando "
                             f"{len(selecionados)}/{len(MODULOS)} módulos: "
                             + ", ".join(m[0] for m in selecionados))
        driver = criar_driver()
        wait = WebDriverWait(driver, 30)

        # Login é crítico: sem ele nenhum módulo consegue rodar.
        janela_original = executar_login(driver, wait)

        for _modulo_id, funcao, nome in selecionados:
            status = executar_modulo(driver, funcao, nome, wait, janela_original)
            resultados.append((nome, status))

        if any(status == STATUS_OK for _, status in resultados):
            mostrar_mensagem_conclusao()
        else:
            mostrar_mensagem_erro()

        adicionar_ao_log("Automação concluída.")

    except Exception as e:
        adicionar_ao_log("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        adicionar_ao_log(f"ERRO CRÍTICO QUE INTERROMPEU A AUTOMAÇÃO: {e}")
        adicionar_ao_log("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        if driver:
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            capturar_screenshot(driver, f"erro_fatal_critico_{timestamp}")
        mostrar_mensagem_erro()
        raise

    finally:
        resumo_execucao(resultados)
        if driver:
            try:
                driver.quit()
                adicionar_ao_log("Driver finalizado.")
            except Exception as e:
                adicionar_ao_log(f"AVISO: falha ao encerrar o driver: {e}")
        adicionar_ao_log("Driver fechado. Automação finalizada.")


if __name__ == "__main__":
    main()
